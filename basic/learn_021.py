import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# 对齐方式
alignment = Alignment(horizontal="center", vertical="center")
# 边框线条
side = Side(color="ff7f50", style="mediumDashed")

# 第一步：创建工作簿（Workbook）
wb = openpyxl.Workbook()

# 第二步：添加工作表（Worksheet）
sheet = wb.active
sheet.title = "期末成绩"
sheet.row_dimensions[1].height = 30
sheet.column_dimensions["E"].width = 120

sheet["E1"] = "平均分"
# 设置字体
sheet.cell(1, 5).font = Font(size=18, bold=True, color="ff1493", name="华文楷体")
# 设置对齐方式
sheet.cell(1, 5).alignment = alignment
# 设置单元格边框
sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
for i in range(2, 7):
    # 公式计算每个学生的平均分
    sheet[f"E{i}"] = f"=average(B{i}:D{i})"
    sheet.cell(i, 5).font = Font(size=12, color="4169e1", italic=True)
    sheet.cell(i, 5).alignment = alignment

titles = ("姓名", "语文", "数学", "英语")
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)

names = ("关羽", "张飞", "赵云", "马超", "黄忠")
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name)
    for col_index in range(2, 5):
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

# 第四步：保存工作簿
wb.save("./res/考试成绩表.xlsx")
